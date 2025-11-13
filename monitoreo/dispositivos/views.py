import logging
import os
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, Http404, HttpResponse
from django.views.decorators.http import require_POST

from django.db.models import Count, Q  
from django.core.paginator import Paginator 
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied, ValidationError
from django.contrib import messages
from django.utils.html import escape
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from usuarios.decorators import cliente_admin_required, cliente_electronico_required, encargado_required

logger = logging.getLogger(__name__)

from .forms import DispositivoForm, ZonaForm, MedicionForm
from .models import Zona, Dispositivo, Medicion, Alerta
from usuarios.models import Organizacion

def get_organizacion_del_usuario(user):
    if not user.is_authenticated:
        return None
    try:
        return user.perfil.organizacion
    except (ObjectDoesNotExist, AttributeError) as e:
        logger.warning(f'Error obteniendo organización del usuario {user.id}: {str(e)}')
        return None

def get_user_role(user):
    if not user.is_authenticated:
        return None
    try:
        return user.perfil.rol
    except (ObjectDoesNotExist, AttributeError) as e:
        logger.warning(f'Error obteniendo rol del usuario {user.id}: {str(e)}')
        return None

def validate_safe_path(path_param):
    """Valida que el parámetro no contenga path traversal"""
    if not path_param:
        return True
    
    dangerous_patterns = ['../', '..\\', '/..', '\\..', '../', '..\\']
    path_str = str(path_param)
    
    for pattern in dangerous_patterns:
        if pattern in path_str:
            return False
    return True

@login_required
def dashboard(request):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)
    
    mediciones_qs = Medicion.objects.select_related('dispositivo')
    zonas_qs = Zona.objects.all()

    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        mediciones_qs = mediciones_qs.filter(dispositivo__zona__organizacion=organizacion_usuario)
        zonas_qs = zonas_qs.filter(organizacion=organizacion_usuario)

    mediciones = mediciones_qs.order_by('-fecha')[:10]
    zonas = zonas_qs.annotate(num_dispositivos=Count('dispositivo'))
    
    alertas_grave = []
    alertas_alta = []
    alertas_media = []

    for medicion in mediciones:
        consumo = medicion.consumo
        if consumo >= 100:
            alertas_grave.append(medicion)
        elif consumo > 80:
            alertas_alta.append(medicion)
        elif consumo > 60:
            alertas_media.append(medicion)

    return render(request, 'dispositivos/panel.html', {
        'mediciones': mediciones,
        'zonas': zonas,
        'alertas_grave': alertas_grave,
        'alertas_alta': alertas_alta,
        'alertas_media': alertas_media,
    })

@login_required
def listar_dispositivos(request):
    try:
        organizacion_usuario = get_organizacion_del_usuario(request.user)
        user_role = get_user_role(request.user)
        
        # Sanitizar parámetros de entrada
        q = escape(request.GET.get('q', '').strip())
        sort = request.GET.get('sort', 'nombre')
        categoria = request.GET.get('categoria')
        page_number = request.GET.get('page')
        size = request.GET.get('size', '10')
        
        # Validar parámetros contra path traversal
        if not all(validate_safe_path(param) for param in [q, sort, categoria, page_number, size]):
            logger.warning(f'Intento de path traversal en listar_dispositivos por usuario {request.user.id}')
            messages.error(request, 'Parámetros inválidos detectados.')
            return redirect('dispositivos:dispositivo_list')

        try:
            page_size = int(size)
            if page_size > 100:  # Limitar tamaño de página
                page_size = 100
        except ValueError:
            page_size = 10

        qs = Dispositivo.objects.select_related('zona').all()
        
        if organizacion_usuario and user_role != 'encargado_ecoenergy':
            qs = qs.filter(zona__organizacion=organizacion_usuario)

        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) |
                Q(categoria__icontains=q) |
                Q(zona__nombre__icontains=q)
            )

        if categoria:
            qs = qs.filter(categoria=categoria)
        
        # Validar campo de ordenamiento
        allowed_sort_fields = ['nombre', '-nombre', 'categoria', '-categoria', 'zona__nombre', '-zona__nombre']
        if sort not in allowed_sort_fields:
            sort = 'nombre'
        
        # Guardar configuración de paginación en sesión
        if 'size' in request.GET:
            request.session['dispositivos_page_size'] = page_size
        elif 'dispositivos_page_size' in request.session:
            page_size = request.session['dispositivos_page_size']
        
        qs = qs.order_by(sort)

        paginator = Paginator(qs, page_size)
        page_obj = paginator.get_page(page_number)

        params = request.GET.copy()
        if 'page' in params:
            del params['page']
        querystring = params.urlencode()
        
        context = {
            'page_obj': page_obj,
            'q': q,
            'sort': sort,
            'categoria': categoria,
            'querystring': querystring,
            'size': str(page_size),
            'categorias': Dispositivo.CATEGORIAS
        }
        
        return render(request, 'dispositivos/dispositivo_list.html', context)
        
    except Exception as e:
        logger.error(f'Error en listar_dispositivos: {str(e)}')
        messages.error(request, 'Error al cargar la lista de dispositivos.')
        return render(request, 'dispositivos/dispositivo_list.html', {'page_obj': None})

@login_required
def detalle_dispositivo(request, dispositivo_id):
    try:
        # Validar que el ID sea seguro
        if not validate_safe_path(dispositivo_id):
            logger.warning(f'Intento de path traversal detectado: {dispositivo_id}')
            raise Http404("Dispositivo no encontrado.")
            
        organizacion_usuario = get_organizacion_del_usuario(request.user)
        user_role = get_user_role(request.user)
        dispositivo = get_object_or_404(Dispositivo, id=dispositivo_id)
        
        if organizacion_usuario and user_role != 'encargado_ecoenergy' and dispositivo.zona.organizacion != organizacion_usuario:
            logger.warning(f'Usuario {request.user.id} intentó acceder a dispositivo {dispositivo_id} sin permisos')
            raise Http404("Dispositivo no encontrado.")

        mediciones = Medicion.objects.filter(dispositivo=dispositivo).order_by('-fecha')
        
        alertas_grave = Alerta.objects.filter(dispositivo=dispositivo, gravedad='Grave').order_by('-fecha')
        alertas_alta = Alerta.objects.filter(dispositivo=dispositivo, gravedad='Alta').order_by('-fecha')
        alertas_media = Alerta.objects.filter(dispositivo=dispositivo, gravedad='Media').order_by('-fecha')

        return render(request, 'dispositivos/dispositivo_detalle.html', {
            'dispositivo': dispositivo,
            'mediciones': mediciones,
            'alertas_grave': alertas_grave,
            'alertas_alta': alertas_alta,
            'alertas_media': alertas_media,
        })
    except Exception as e:
        logger.error(f'Error en detalle_dispositivo: {str(e)}')
        raise Http404("Error al cargar el dispositivo.")

@login_required
def crear_dispositivo(request):
    # Permitir a todos los usuarios autenticados crear dispositivos
    if request.method == "POST":
        form = DispositivoForm(request.POST, user=request.user)
        if form.is_valid():
            zona_seleccionada = form.cleaned_data['zona']
            organizacion_usuario = get_organizacion_del_usuario(request.user)
            user_role = get_user_role(request.user)
            
            if organizacion_usuario and user_role != 'encargado_ecoenergy' and zona_seleccionada.organizacion != organizacion_usuario:
                form.add_error('zona', 'Esta zona no pertenece a tu organización.')
            else:
                dispositivo = form.save()
                messages.success(request, f'Dispositivo "{dispositivo.nombre}" creado exitosamente.')
                return redirect("dispositivos:dispositivo_list")
    else:
        form = DispositivoForm(user=request.user)
    
    return render(request, "dispositivos/dispositivo_form.html", {"form": form})

@login_required
def editar_dispositivo(request, dispositivo_id):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)
    dispositivo = get_object_or_404(Dispositivo, id=dispositivo_id)

    if organizacion_usuario and user_role != 'encargado_ecoenergy' and dispositivo.zona.organizacion != organizacion_usuario:
        raise Http404("Dispositivo no encontrado.")

    if request.method == "POST":
        form = DispositivoForm(request.POST, instance=dispositivo, user=request.user)
        if form.is_valid():
            zona_seleccionada = form.cleaned_data['zona']
            if organizacion_usuario and user_role != 'encargado_ecoenergy' and zona_seleccionada.organizacion != organizacion_usuario:
                form.add_error('zona', 'Esta zona no pertenece a tu organización.')
            else:
                dispositivo = form.save()
                messages.success(request, f'Dispositivo "{dispositivo.nombre}" actualizado exitosamente.')
                return redirect("dispositivos:dispositivo_detail", dispositivo_id=dispositivo.id)
    else:
        form = DispositivoForm(instance=dispositivo, user=request.user)
    
    return render(request, "dispositivos/dispositivo_form.html", {"form": form})

@login_required
@cliente_admin_required
@require_POST
def eliminar_dispositivo(request, dispositivo_id):
    try:
        # Validar path traversal
        if not validate_safe_path(dispositivo_id):
            logger.warning(f'Intento de path traversal en eliminar_dispositivo: {dispositivo_id}')
            return JsonResponse({"ok": False, "message": "ID inválido."}, status=400)
            
        organizacion_usuario = get_organizacion_del_usuario(request.user)
        user_role = get_user_role(request.user)
        dispositivo = get_object_or_404(Dispositivo, id=dispositivo_id)

        if organizacion_usuario and user_role != 'encargado_ecoenergy' and dispositivo.zona.organizacion != organizacion_usuario:
            logger.warning(f'Usuario {request.user.id} intentó eliminar dispositivo {dispositivo_id} sin permisos')
            return JsonResponse({"ok": False, "message": "Permiso denegado."}, status=403)

        nombre = dispositivo.nombre
        dispositivo.delete()
        logger.info(f'Dispositivo {nombre} eliminado por usuario {request.user.id}')
        return JsonResponse({"ok": True, "message": f"Dispositivo '{nombre}' eliminado"})
        
    except Exception as e:
        logger.error(f'Error eliminando dispositivo {dispositivo_id}: {str(e)}')
        return JsonResponse({"ok": False, "message": "Error interno del servidor."}, status=500)

@login_required
def listar_zonas(request):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)
    
    zonas_qs = Zona.objects.all()
    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        zonas_qs = zonas_qs.filter(organizacion=organizacion_usuario)
    
    zonas = zonas_qs.annotate(num_dispositivos=Count('dispositivo')).order_by('nombre')
    
    return render(request, 'dispositivos/zona_list.html', {'zonas': zonas})

@login_required
def crear_zona(request):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)

    if request.method == 'POST':
        form = ZonaForm(request.POST)
        if form.is_valid():
            zona = form.save(commit=False)
            if user_role == 'encargado_ecoenergy':
                # El encargado puede crear zonas para cualquier organización
                # Por ahora asignamos una por defecto
                organizacion, created = Organizacion.objects.get_or_create(
                    nombre='EcoEnergy Default'
                )
                zona.organizacion = organizacion
            elif organizacion_usuario:
                zona.organizacion = organizacion_usuario
            else:
                messages.error(request, 'No tienes una organización asignada.')
                return render(request, "dispositivos/zona_form.html", {"form": form})
            
            zona.save()
            messages.success(request, f'Zona "{zona.nombre}" creada exitosamente.')
            return redirect('dispositivos:zona_list')
    else:
        form = ZonaForm()
        
    return render(request, "dispositivos/zona_form.html", {"form": form})

@login_required
@cliente_admin_required
def editar_zona(request, zona_id):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)
    
    qs = Zona.objects.all()
    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        qs = qs.filter(organizacion=organizacion_usuario) 
        
    zona = get_object_or_404(qs, id=zona_id)
    
    if request.method == 'POST':
        form = ZonaForm(request.POST, instance=zona)
        if form.is_valid():
            form.save()
            messages.success(request, f'Zona "{zona.nombre}" actualizada exitosamente.')
            return redirect('dispositivos:zona_list')
    else:
        form = ZonaForm(instance=zona)
        
    return render(request, "dispositivos/zona_form.html", {"form": form, "zona": zona})

@login_required
@cliente_admin_required
@require_POST
def eliminar_zona(request, zona_id):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)
    
    qs = Zona.objects.all()
    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        qs = qs.filter(organizacion=organizacion_usuario) 
        
    zona = get_object_or_404(qs, id=zona_id)
    
    if zona.dispositivo_set.count() > 0:
        return JsonResponse({"ok": False, "message": "No se puede eliminar la zona porque tiene dispositivos asociados."}, status=400)
    
    try:
        nombre = zona.nombre
        zona.delete()
        return JsonResponse({"ok": True, "message": f"Zona '{nombre}' eliminada"})
    except Exception as e:
        return JsonResponse({"ok": False, "message": str(e)}, status=400)

@login_required
def listar_mediciones(request):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)

    dispositivo_id = request.GET.get('dispositivo_id', '')
    fecha_inicio = request.GET.get('fecha_inicio', '')
    fecha_fin = request.GET.get('fecha_fin', '')
    page_number = request.GET.get('page') 
    size = request.GET.get('size', '10')

    try:
        page_size = int(size)
    except ValueError:
        page_size = 10

    mediciones_qs = Medicion.objects.select_related('dispositivo', 'dispositivo__zona')
    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        mediciones_qs = mediciones_qs.filter(dispositivo__zona__organizacion=organizacion_usuario)

    dispositivos_para_filtro = Dispositivo.objects.order_by('nombre')
    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        dispositivos_para_filtro = dispositivos_para_filtro.filter(zona__organizacion=organizacion_usuario)

    if dispositivo_id:
        mediciones_qs = mediciones_qs.filter(dispositivo_id=dispositivo_id)
    
    if fecha_inicio:
        try:
            mediciones_qs = mediciones_qs.filter(fecha__date__gte=fecha_inicio)
        except:
            fecha_inicio = ''
    
    if fecha_fin:
        try:
            mediciones_qs = mediciones_qs.filter(fecha__date__lte=fecha_fin)
        except:
            fecha_fin = ''

    mediciones_qs = mediciones_qs.order_by('-fecha')
    paginator = Paginator(mediciones_qs, page_size) 
    page_obj = paginator.get_page(page_number) 

    params = request.GET.copy()
    if 'page' in params:
        del params['page']
    querystring = params.urlencode() 

    context = {
        'page_obj': page_obj,
        'querystring': querystring,
        'dispositivos_para_filtro': dispositivos_para_filtro,
        'dispositivo_id_seleccionado': dispositivo_id,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'size': size
    }
    
    return render(request, 'dispositivos/mediciones_list.html', context)

@login_required
def crear_medicion(request):
    if request.method == 'POST':
        form = MedicionForm(request.POST, user=request.user)
        if form.is_valid():
            medicion = form.save()
            messages.success(request, f'Medición creada exitosamente para {medicion.dispositivo.nombre}.')
            return redirect('dispositivos:medicion_list')
    else:
        form = MedicionForm(user=request.user)
    
    return render(request, 'dispositivos/medicion_form.html', {'form': form})

@login_required
@cliente_electronico_required
def editar_medicion(request, medicion_id):
    medicion = get_object_or_404(Medicion, id=medicion_id)
    
    if request.method == 'POST':
        form = MedicionForm(request.POST, instance=medicion, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Medición actualizada exitosamente.')
            return redirect('dispositivos:medicion_list')
    else:
        form = MedicionForm(instance=medicion, user=request.user)
    
    return render(request, 'dispositivos/medicion_form.html', {'form': form, 'medicion': medicion})

@login_required
@cliente_electronico_required
@require_POST
def eliminar_medicion(request, medicion_id):
    medicion = get_object_or_404(Medicion, id=medicion_id)
    
    try:
        dispositivo_nombre = medicion.dispositivo.nombre
        medicion.delete()
        return JsonResponse({"ok": True, "message": f"Medición de {dispositivo_nombre} eliminada"})
    except Exception as e:
        return JsonResponse({"ok": False, "message": str(e)}, status=400)

@login_required
def detalle_medicion(request, medicion_id):
    medicion = get_object_or_404(Medicion, id=medicion_id)
    return render(request, 'dispositivos/medicion_detalle.html', {'medicion': medicion})

@login_required
def exportar_dispositivos_excel(request):
    try:
        organizacion_usuario = get_organizacion_del_usuario(request.user)
        user_role = get_user_role(request.user)
        
        # Sanitizar parámetros
        q = escape(request.GET.get('q', '').strip())
        categoria = request.GET.get('categoria')
        
        # Validar path traversal
        if not all(validate_safe_path(param) for param in [q, categoria]):
            logger.warning(f'Intento de path traversal en exportar_dispositivos_excel por usuario {request.user.id}')
            return HttpResponse('Parámetros inválidos', status=400)
        
        qs = Dispositivo.objects.select_related('zona').all()
        
        if organizacion_usuario and user_role != 'encargado_ecoenergy':
            qs = qs.filter(zona__organizacion=organizacion_usuario)
        
        if q:
            qs = qs.filter(
                Q(nombre__icontains=q) |
                Q(categoria__icontains=q) |
                Q(zona__nombre__icontains=q)
            )
        
        if categoria:
            qs = qs.filter(categoria=categoria)
        
        # Limitar resultados para evitar sobrecarga
        qs = qs.order_by('nombre')[:1000]
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Dispositivos"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        headers = ['ID', 'Nombre', 'Categoría', 'Zona', 'Organización', 'Watts']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        for row, dispositivo in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=dispositivo.id)
            ws.cell(row=row, column=2, value=escape(str(dispositivo.nombre)))
            ws.cell(row=row, column=3, value=escape(str(dispositivo.categoria)))
            ws.cell(row=row, column=4, value=escape(str(dispositivo.zona.nombre)) if dispositivo.zona else 'Sin zona')
            ws.cell(row=row, column=5, value=escape(str(dispositivo.zona.organizacion.nombre)) if dispositivo.zona else 'Sin organización')
            ws.cell(row=row, column=6, value=dispositivo.watts)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Sanitizar nombre de archivo
        filename = "dispositivos.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        logger.info(f'Exportación de dispositivos realizada por usuario {request.user.id}')
        return response
        
    except Exception as e:
        logger.error(f'Error en exportar_dispositivos_excel: {str(e)}')
        return HttpResponse('Error al generar el archivo', status=500)

@login_required
def listar_alertas(request):
    organizacion_usuario = get_organizacion_del_usuario(request.user)
    user_role = get_user_role(request.user)

    dispositivo_id = request.GET.get('dispositivo_id', '').strip()
    gravedad = request.GET.get('gravedad', '').strip()
    fecha_inicio = request.GET.get('fecha_inicio', '').strip()
    fecha_fin = request.GET.get('fecha_fin', '').strip()
    page_number = request.GET.get('page') 
    size = request.GET.get('size', '10')

    try:
        page_size = int(size)
    except ValueError:
        page_size = 10
    
    alertas_qs = Alerta.objects.select_related('dispositivo', 'dispositivo__zona')
    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        alertas_qs = alertas_qs.filter(dispositivo__zona__organizacion=organizacion_usuario)

    dispositivos_para_filtro = Dispositivo.objects.order_by('nombre')
    if organizacion_usuario and user_role != 'encargado_ecoenergy':
        dispositivos_para_filtro = dispositivos_para_filtro.filter(zona__organizacion=organizacion_usuario)

    if dispositivo_id:
        alertas_qs = alertas_qs.filter(dispositivo_id=dispositivo_id)
    
    # Solo filtrar por gravedad si no es vacío (permite mostrar todas cuando se selecciona "Todas las gravedades")
    if gravedad:
        alertas_qs = alertas_qs.filter(gravedad=gravedad)
    
    if fecha_inicio:
        try:
            alertas_qs = alertas_qs.filter(fecha__date__gte=fecha_inicio)
        except:
            fecha_inicio = ''
    
    if fecha_fin:
        try:
            alertas_qs = alertas_qs.filter(fecha__date__lte=fecha_fin)
        except:
            fecha_fin = ''

    alertas_qs = alertas_qs.order_by('-fecha')
    paginator = Paginator(alertas_qs, page_size) 
    page_obj = paginator.get_page(page_number) 

    # Construir querystring limpio: solo incluir parámetros que no estén vacíos
    params = request.GET.copy()
    if 'page' in params:
        del params['page']
    # Remover parámetros con valores vacíos para evitar &gravedad=&... en URLs de paginación
    clean_params = [(k, v) for k, v in params.items() if v and str(v).strip()]
    querystring = '&'.join([f'{k}={v}' for k, v in clean_params]) if clean_params else '' 

    context = {
        'page_obj': page_obj,
        'querystring': querystring,
        'dispositivos_para_filtro': dispositivos_para_filtro,
        'dispositivo_id_seleccionado': dispositivo_id,
        'gravedad_seleccionada': gravedad,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'size': size,
        'GRAVEDAD_CHOICES': Alerta.GRAVEDAD_CHOICES
    }
    
    return render(request, 'dispositivos/alertas_list.html', context)

def custom_404(request, exception):
    logger.warning(f'Página no encontrada: {request.path} por usuario {getattr(request.user, "id", "anónimo")}')
    return render(request, '404.html', status=404)